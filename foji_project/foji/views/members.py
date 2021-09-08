# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile

from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils.decorators import method_decorator
from django import forms

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from ..models.personal_information import PersonalInformation
from ..models.orchestra import Orchestra
from ..models.director import Director
from ..models.instructor import Instructor
from ..models.cast_member import CastMember


EXCEL_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

CAST_MEMBER_HEADER_CELLS = (
    'Nombre',
    'Apellido',
    'Instrumento',
    'Genero',
    'Fecha de nacimiento',
    'Email',
    'Teléfono',
    'RUT',
)

FILENAME = 'integrantes.xlsx'

INSTRUMENT_LIST = '"Violín,Viola,Cello,Contrabajo,Flautín,Flauta,Oboe,Corno Inglés,Clarinete,Fagot,Contrafagot,Saxo Soprano,Saxo Alto,Saxo Tenor,Saxo Barítono,Corno Francés,Trompeta,Trombón,Tuba,Piano,Arpa,Percusión,Guitarra"' 

GENDER_LIST = '"M,F"'


class DirectorForm(forms.Form):
    first_name = forms.CharField()
    last_name = forms.CharField()
    gender = forms.CharField()
    birth_date= forms.DateField()
    email = forms.EmailField()
    phone_number_mobile = forms.CharField()
    social_id = forms.CharField()


class InstructorForm(forms.Form):
    first_name = forms.CharField()
    last_name = forms.CharField()
    gender = forms.CharField()
    birth_date= forms.DateField()
    email = forms.EmailField()
    phone_number_mobile = forms.CharField()
    social_id = forms.CharField()


class CastMemberForm(forms.Form):
    first_name = forms.CharField()
    last_name = forms.CharField()
    gender = forms.CharField()
    birth_date= forms.DateField()
    email = forms.EmailField()
    phone_number_mobile = forms.CharField()
    social_id = forms.CharField()


def create_row(member):
    try:
        first_name = member.personal_information.first_name
    except:
        first_name = None

    try:
        last_name = member.personal_information.last_name
    except:
        last_name = None

    try:
        gender = member.personal_information.gender
    except:
        gender = None

    try:
        birth_date = member.personal_information.birth_date
    except:
        birth_date = None

    try:
        email = member.personal_information.email
    except:
        email = None

    try:
        phone_number_mobile = member.personal_information.phone_number_mobile
    except:
        phone_number_mobile = None

    try:
        social_id = member.personal_information.social_id
    except:
        social_id = None


    result = [
        first_name,
        last_name,
        None,
        gender,
        birth_date,
        email,
        phone_number_mobile,
        social_id,
    ]

    return result


@method_decorator(csrf_exempt, name='dispatch')
class CastMembersExcelView(View):
    INTRODUCTION_TEXT = """INSTRUCCIONES:

    Navega por las pestañas "Director", "Instructores" y
    "Elenco" en este documento y rellena todos los campos
    con la información de tu orquesta.

    Es importante rellenar todos los campos ya que son
    requeridos."""

    def get(self, request, pk):
        self.orchestra = Orchestra.objects.get(pk=pk)

        wb = Workbook()

        sheet_introduction = wb.active
        sheet_introduction.title = 'Introducción'

        sheet_directors = wb.create_sheet('Director')
        sheet_instructors = wb.create_sheet('Instructores')
        sheet_cast = wb.create_sheet('Elenco')

        # Introduction sheet
        # TODO: Add introduction.
        sheet_introduction.merge_cells('A1:Z20')
        sheet_introduction['A1'].alignment = Alignment(wrap_text=True)
        sheet_introduction['A1'] = self.INTRODUCTION_TEXT


        # Director sheet
        sheet_directors.append(CAST_MEMBER_HEADER_CELLS)

        dv = DataValidation(
            type='list',
            formula1=INSTRUMENT_LIST,
            allow_blank=True,
        )
        # Instruments
        dv.error = 'Elige un instrumento válido.'
        dv.errorTitle = 'Instrumento inválido'
        dv.prompt = 'Selecciona un instrumento.'
        dv.promptTitle = 'Lista de instrumentos'
        dv.add('C2:C100')
        sheet_directors.add_data_validation(dv)

        # Gender
        dv = DataValidation(
            type='list',
            formula1=GENDER_LIST,
            allow_blank=True,
        )
        dv.error = 'Elige un genero válido.'
        dv.errorTitle = 'Genero inválido'
        dv.prompt = 'Selecciona un genero.'
        dv.promptTitle = 'Lista de generos'
        dv.add('D2:D100')
        sheet_directors.add_data_validation(dv)


        # Instructor sheet
        sheet_instructors.append(CAST_MEMBER_HEADER_CELLS)

        # Instruments
        dv = DataValidation(
            type='list',
            formula1=INSTRUMENT_LIST,
            allow_blank=True,
        )
        dv.error = 'Elige un instrumento válido.'
        dv.errorTitle = 'Instrumento inválido'
        dv.prompt = 'Selecciona un instrumento.'
        dv.promptTitle = 'Lista de instrumentos'
        dv.add('C2:C100')
        sheet_instructors.add_data_validation(dv)

        # Gender
        dv = DataValidation(
            type='list',
            formula1=GENDER_LIST,
            allow_blank=True,
        )
        dv.error = 'Elige un genero válido.'
        dv.errorTitle = 'Genero inválido'
        dv.prompt = 'Selecciona un genero.'
        dv.promptTitle = 'Lista de generos'
        dv.add('D2:D100')
        sheet_instructors.add_data_validation(dv)

        # Birth date

        # Cast sheet
        sheet_cast.append(CAST_MEMBER_HEADER_CELLS)

        dv = DataValidation(
            type='list',
            formula1=INSTRUMENT_LIST,
            allow_blank=True,
        )

        # Instruments
        dv.error = 'Elige un instrumento válido.'
        dv.errorTitle = 'Instrumento inválido'
        dv.prompt = 'Selecciona un instrumento.'
        dv.promptTitle = 'Lista de instrumentos'
        dv.add('C2:C100')
        sheet_cast.add_data_validation(dv)

        # Gender
        dv = DataValidation(
            type='list',
            formula1=GENDER_LIST,
            allow_blank=True,
        )
        dv.error = 'Elige un genero válido.'
        dv.errorTitle = 'Genero inválido'
        dv.prompt = 'Selecciona un genero.'
        dv.promptTitle = 'Lista de generos'
        dv.add('D2:D100')
        sheet_cast.add_data_validation(dv)

        # Birth date

        ##

        if self.orchestra.director:
            sheet_directors.append(create_row(self.orchestra.director))

        for instructor in self.orchestra.instructors.all():
            sheet_instructors.append(create_row(instructor))

        for cast_member in self.orchestra.cast_members.all():
            sheet_cast.append(create_row(cast_member))

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content_type=EXCEL_MIME)
        response['Content-Disposition'] = 'attachment; filename={}'.format(FILENAME)
        response.write(stream)

        return response

    def post(self, request, pk):
        self.orchestra = Orchestra.objects.get(pk=pk)
        file_object = request.FILES['file']

        wb = load_workbook(file_object)

        sheet_director = wb['Director']
        sheet_instructors = wb['Instructores']
        sheet_cast = wb['Elenco']

        self.create_directors(sheet_director)
        self.create_instructors(sheet_instructors)
        self.create_cast(sheet_cast)

        self.orchestra.save()

        return redirect('/orquesta/{}/'.format(pk))


    def create_directors(self, sheet):
        print(sheet.max_row)
        try:
            self.orchestra.director.delete()
        except:
            pass

        for row in sheet.iter_rows(min_row=2):
            data = {}

            data['first_name'] = row[0].value
            data['last_name'] = row[1].value
            data['instrument'] = row[2].value
            data['gender'] = row[3].value
            data['birth_date'] = row[4].value
            data['email'] = row[5].value
            data['phone_number_mobile'] = row[6].value
            data['social_id'] = row[7].value

            form = DirectorForm(data)

            if form.is_valid():
                try:
                    data = dict(form.cleaned_data)
                    personal_information = PersonalInformation.objects.create(
                        first_name=data['first_name'],
                        last_name=data['last_name'],
                        gender=data['gender'],
                        phone_number_mobile=data['phone_number_mobile'],
                        birth_date=data['birth_date'],
                        email=data['email'],
                        social_id=data['social_id'],
                    )

                    director = Director.objects.create(
                        personal_information=personal_information,
                    )

                    self.orchestra.director = director

                except Exception as e:
                    print(e)
                    continue


    def create_instructors(self, sheet):
        print(sheet.max_row)
        self.orchestra.instructors.all().delete()
        for row in sheet.iter_rows(min_row=2):
            data = {}

            try:
                data['first_name'] = row[0].value
                data['last_name'] = row[1].value
                data['instrument'] = row[2].value
                data['gender'] = row[3].value
                data['birth_date'] = row[4].value
                data['email'] = row[5].value
                data['phone_number_mobile'] = row[6].value
                data['social_id'] = row[7].value
            except:
                continue

            form = InstructorForm(data)

            if form.is_valid():
                try:
                    data = dict(form.cleaned_data)
                    personal_information = PersonalInformation.objects.create(
                        first_name=data['first_name'],
                        last_name=data['last_name'],
                        gender=data['gender'],
                        phone_number_mobile=data['phone_number_mobile'],
                        birth_date=data['birth_date'],
                        email=data['email'],
                        social_id=data['social_id'],
                    )

                    instructor = Instructor.objects.create(
                        personal_information=personal_information,
                        orchestra=self.orchestra,
                    )
                except Exception as e:
                    print(e)
                    continue

    def create_cast(self, sheet):
        print(sheet.max_row)
        self.orchestra.cast_members.all().delete()
        for row in sheet.iter_rows(min_row=2):
            data = {}

            try:
                data['first_name'] = row[0].value
                data['last_name'] = row[1].value
                data['instrument'] = row[2].value
                data['gender'] = row[3].value
                data['birth_date'] = row[4].value
                data['email'] = row[5].value
                data['phone_number_mobile'] = row[6].value
                data['social_id'] = row[7].value
            except:
                continue

            form = CastMemberForm(data)

            if form.is_valid():
                try:
                    data = dict(form.cleaned_data)
                    personal_information = PersonalInformation.objects.create(
                        first_name=data['first_name'],
                        last_name=data['last_name'],
                        gender=data['gender'],
                        phone_number_mobile=data['phone_number_mobile'],
                        birth_date=data['birth_date'],
                        email=data['email'],
                        social_id=data['social_id'],
                    )

                    cast_member = CastMember.objects.create(
                            personal_information=personal_information,
                            orchestra=self.orchestra,
                    )
                except Exception as e:
                    print(e)
                    continue
