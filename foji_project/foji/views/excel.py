# -*- coding: utf-8 -*-
from tempfile import NamedTemporaryFile

from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils.decorators import method_decorator
from django.utils import timezone
from django import forms

from openpyxl import Workbook, load_workbook

from ..models.orchestra import Orchestra
from ..models.coordinator import Coordinator
from ..models.administrator import Administrator
from ..models.director import Director
from ..models.instructor import Instructor
from ..models.cast_member import CastMember


FILENAME = 'result.xlsx'
EXCEL_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class OrchestraExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.active

        ws.append([
            'Nombre',
            'Tipo',
            'Region',
            'Provincia',
            'Comuna',
            'Ciudad',
            'Fecha de creacion',
            'Fecha de registro',
            'Ultima modificacion',
            'Coordinador',
            'Email coordinador',
            'Institucion sostenedora',
            'Email institucion sostenedora',
            'Representante legal',
            'Email representante legal',
            'Director',
            'Email director',
            'Integrantes',
        ])

        orchestras = self.filter_results(request)

        for orchestra in orchestras:
            try:
                region = orchestra.area.province.region.name
            except:
                region = ''

            try:
                province = orchestra.area.province.name
            except:
                province = ''

            try:
                area = orchestra.area.name
            except:
                area = ''

            try:
                coordinator_name = orchestra.coordinator.personal_information.full_name()
            except:
                coordinator_name = ''

            try:
                coordinator_email = orchestra.coordinator.user.username
            except:
                coordinator_email = ''

            try:
                institution_name = orchestra.institution.name
            except:
                institution_name = ''

            try:
                institution_email = orchestra.institution.email
            except:
                institution_email = ''

            try:
                legal_representation_name = \
                    orchestra.institution.legal_representation.full_name()
            except:
                legal_representation_name = ''

            try:
                legal_representation_email = \
                    orchestra.institution.legal_representation.email
            except:
                legal_representation_email = ''

            try:
                director_name = orchestra.director.personal_information.full_name()
            except:
                director_name = ''

            try:
                director_email = orchestra.director.personal_information.email
            except:
                director_email = ''

            try:
                members_count = orchestra.members_count()
            except:
                members_count = ''

            ws.append([
                orchestra.name,
                orchestra.orchestra_type,
                region,
                province,
                area,
                orchestra.city,
                orchestra.creation_date,
                orchestra.created_date,
                orchestra.modified_date,
                coordinator_name,
                coordinator_email,
                institution_name,
                institution_email,
                legal_representation_name,
                legal_representation_email,
                director_name,
                director_email,
                members_count,
            ])

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content_type=EXCEL_MIME)
        response['Content-Disposition'] = 'attachment; filename={}'.format(FILENAME)
        response.write(stream)

        return response

    def filter_results(self, request):
        queryset = Orchestra.objects.all()

        code = request.GET.get('code', None)
        name = request.GET.get('name', None)
        region = request.GET.get('region', None)
        province = request.GET.get('province', None)
        area = request.GET.get('area', None)
        city = request.GET.get('city', None)
        active = request.GET.get('active', None)
        orchestra_type = request.GET.get('orchestra_type', None)

        if code is not None:
            splitted = code.split('-')
            r = splitted[0]

            queryset = queryset.filter(
                area__province__region__code__icontains=r,
            )

            try:
                a = splitted[1]
                queryset = queryset.filter(
                    area__code__icontains=a,
                )
            except:
                pass

            try:
                i = int(splitted[2])
                queryset = queryset.filter(
                    id=i,
                )
            except:
                pass

        if name is not None:
            queryset = queryset.filter(
                name__icontains=name,
            )

        if region is not None:
            queryset = queryset.filter(
                area__province__region__id=int(region),
            )

        if province is not None:
            queryset = queryset.filter(
                area__province__id=int(province),
            )

        if area is not None:
            queryset = queryset.filter(
                area__id=int(area),
            )

        if city is not None:
            queryset = queryset.filter(
                city__icontains=city,
            )

        if active is not None:
            if active == 'true':
                active = True
            else:
                active = False

            queryset = queryset.filter(
                is_active=active,
            )

        if orchestra_type is not None:
            queryset = queryset.filter(
                orchestra_type=orchestra_type,
            )

        return queryset



class CoordinatorExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.active

        ws.append([
            'Nombre',
            'Apellido',
            'RUT',
            'Email',
            'Teléfono',
            'Fecha de nacimiento',
            'Edad',
            'Fecha de registro',
            'Número de orquestas',
        ])

        print('OK')
        print(request.GET)

        coordinators = self.filter_results(request)

        for coordinator in coordinators:
            try:
                first_name = coordinator.personal_information.first_name
            except:
                first_name = ''

            try:
                last_name = coordinator.personal_information.last_name
            except:
                last_name = ''

            try:
                social_id = coordinator.personal_information.social_id
            except:
                social_id = ''

            try:
                phone_number_mobile = coordinator.personal_information.phone_number_mobile
            except:
                phone_number_mobile = ''

            try:
                birth_date = coordinator.personal_information.birth_date
            except:
                birth_date = ''

            try:
                age = timezone.now().year - birth_date.year
            except:
                age = ''


            ws.append([
                first_name,
                last_name,
                social_id,
                coordinator.user.username,
                phone_number_mobile,
                birth_date,
                age,
                coordinator.created_date,
                coordinator.orchestras.all().count(),
            ])


        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content_type=EXCEL_MIME)
        response['Content-Disposition'] = 'attachment; filename={}'.format(FILENAME)
        response.write(stream)

        return response


    def filter_results(self, request):
        print('Ready to filter')
        queryset = Coordinator.objects.all()

        first_name = request.GET.get('first_name', None)
        last_name = request.GET.get('last_name', None)
        social_id = request.GET.get('social_id', None)
        email = request.GET.get('email', None)
        phone_number_mobile = request.GET.get('phone_number_mobile', None)
        region = request.GET.get('region', None)
        province = request.GET.get('province', None)
        area = request.GET.get('area', None)

        print('firstname', first_name)
        print('email', email)

        if first_name is not None:
            queryset = queryset.filter(
                personal_information__first_name__icontains=first_name,
            )

        if last_name is not None:
            queryset = queryset.filter(
                personal_information__last_name__icontains=last_name,
            )

        if social_id is not None:
            queryset = queryset.filter(
                personal_information__social_id__icontains=social_id,
            )

        if email is not None:
            queryset = queryset.filter(
                user__username__icontains=email,
            )

        if phone_number_mobile is not None:
            queryset = queryset.filter(
                personal_information__phone_number_mobile__icontains=phone_number_mobile,
            )

        if region is not None:
            queryset = queryset.filter(
                orchestra__area__province__region_id=int(region),
            )

        if province is not None:
            queryset = queryset.filter(
                orchestra__area__province__id=int(province),
            )

        if area is not None:
            queryset = queryset.filter(
                orchestra__area__id=int(area),
            )

        return queryset


class AdministratorExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.active

        ws.append([
            'Nombre',
            'Apellido',
            'Email',
            'Fecha de registro',
        ])

        administrators = self.filter_results(request)

        for administrator in administrators:
            try:
                first_name = administrator.personal_information.first_name
            except:
                first_name = ''

            try:
                last_name = administrator.personal_information.last_name
            except:
                last_name = ''

            try:
                email = administrator.personal_information.email
            except:
                email = ''

            ws.append([
                first_name,
                last_name,
                email,
                administrator.created_date,
            ])


        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content_type=EXCEL_MIME)
        response['Content-Disposition'] = 'attachment; filename={}'.format(FILENAME)
        response.write(stream)

        return response


    def filter_results(self, request):
        queryset = Administrator.objects.all()

        first_name = request.GET.get('first_name', None)
        last_name = request.GET.get('last_name', None)
        email = request.GET.get('email', None)

        if first_name is not None:
            queryset = queryset.filter(
                personal_information__first_name__icontains=first_name,
            )

        if last_name is not None:
            queryset = queryset.filter(
                personal_information__last_name__icontains=last_name,
            )

        if email is not None:
            queryset = queryset.filter(
                personal_information__email__icontains=email,
            )

        return queryset


class DirectorExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.active

        ws.append([
            'Nombre',
            'Apellido',
            'RUT',
            'Email',
            'Teléfono',
            'Fecha de nacimiento',
            'Edad',
            'Orquesta',
        ])

        directors = self.filter_results(request)

        for director in directors:
            try:
                orchestra = director.orchestras.all().first().name
            except:
                orchestra = ''

            try:
                first_name = director.personal_information.first_name
            except:
                first_name = ''

            try:
                last_name = director.personal_information.last_name
            except:
                last_name = ''

            try:
                social_id = director.personal_information.social_id
            except:
                social_id = ''

            try:
                email = director.personal_information.email
            except:
                email = ''

            try:
                phone_number_mobile = director.personal_information.phone_number_mobile
            except:
                phone_number_mobile = ''

            try:
                birth_date = director.personal_information.birth_date
            except:
                birth_date = ''

            try:
                age = timezone.now().year - birth_date.year
            except:
                age = ''

            ws.append([
                first_name,
                last_name,
                social_id,
                email,
                phone_number_mobile,
                birth_date,
                age,
                orchestra,
            ])


        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content_type=EXCEL_MIME)
        response['Content-Disposition'] = 'attachment; filename={}'.format(FILENAME)
        response.write(stream)

        return response


    def filter_results(self, request):
        queryset = Director.objects.all()

        first_name = request.GET.get('first_name', None)
        last_name = request.GET.get('last_name', None)
        social_id = request.GET.get('social_id', None)
        email = request.GET.get('email', None)
        phone_number_mobile = request.GET.get('phone_number_mobile', None)
        region = request.GET.get('region', None)
        province = request.GET.get('province', None)
        area = request.GET.get('area', None)

        if first_name is not None:
            queryset = queryset.filter(
                personal_information__first_name__icontains=first_name,
            )

        if last_name is not None:
            queryset = queryset.filter(
                personal_information__last_name__icontains=last_name,
            )

        if social_id is not None:
            queryset = queryset.filter(
                personal_information__social_id__icontains=social_id,
            )

        if email is not None:
            queryset = queryset.filter(
                personal_information__email__icontains=email,
            )

        if phone_number_mobile is not None:
            queryset = queryset.filter(
                personal_information__phone_number_mobile__icontains=phone_number_mobile,
            )

        if region is not None:
            queryset = queryset.filter(
                orchestras__area__province__region_id=int(region),
            )

        if province is not None:
            queryset = queryset.filter(
                orchestras__area__province__id=int(province),
            )

        if area is not None:
            queryset = queryset.filter(
                orchestras__area__id=int(area),
            )

        return queryset


class InstructorExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.active

        ws.append([
            'Nombre',
            'Apellido',
            'RUT',
            'Email',
            'Teléfono',
            'Fecha de nacimiento',
            'Edad',
            'Instrumento',
            'Estudiantes',
            'Orquesta',
        ])

        instructors = self.filter_results(request)

        for instructor in instructors:
            try:
                orchestra = instructor.orchestra.name
            except:
                orchestra = ''

            try:
                first_name = instructor.personal_information.first_name
            except:
                first_name = ''

            try:
                last_name = instructor.personal_information.last_name
            except:
                last_name = ''

            try:
                social_id = instructor.personal_information.social_id
            except:
                social_id = ''

            try:
                email = instructor.personal_information.email
            except:
                email = ''

            try:
                phone_number_mobile = instructor.personal_information.phone_number_mobile
            except:
                phone_number_mobile = ''

            try:
                birth_date = instructor.personal_information.birth_date
            except:
                birth_date = ''

            try:
                age = timezone.now().year - birth_date.year
            except:
                age = ''

            ws.append([
                first_name,
                last_name,
                social_id,
                email,
                phone_number_mobile,
                birth_date,
                age,
                instructor.instrument,
                instructor.students,
                orchestra,
            ])


        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content_type=EXCEL_MIME)
        response['Content-Disposition'] = 'attachment; filename={}'.format(FILENAME)
        response.write(stream)

        return response


    def filter_results(self, request):
        queryset = Instructor.objects.all()

        first_name = request.GET.get('first_name', None)
        last_name = request.GET.get('last_name', None)
        social_id = request.GET.get('social_id', None)
        email = request.GET.get('email', None)
        phone_number_mobile = request.GET.get('phone_number_mobile', None)
        region = request.GET.get('region', None)
        province = request.GET.get('province', None)
        area = request.GET.get('area', None)
        instrument = request.GET.get('instrument', None)

        if first_name is not None:
            queryset = queryset.filter(
                personal_information__first_name__icontains=first_name,
            )

        if last_name is not None:
            queryset = queryset.filter(
                personal_information__last_name__icontains=last_name,
            )

        if social_id is not None:
            queryset = queryset.filter(
                personal_information__social_id__icontains=social_id,
            )

        if email is not None:
            queryset = queryset.filter(
                personal_information__email__icontains=email,
            )

        if phone_number_mobile is not None:
            queryset = queryset.filter(
                personal_information__phone_number_mobile__icontains=phone_number_mobile,
            )

        if region is not None:
            queryset = queryset.filter(
                orchestra__area__province__region_id=int(region),
            )

        if province is not None:
            queryset = queryset.filter(
                orchestra__area__province__id=int(province),
            )

        if area is not None:
            queryset = queryset.filter(
                orchestra__area__id=int(area),
            )

        if instrument is not None:
            queryset = queryset.filter(
                instrument__icontains=instrument,
            )

        return queryset


class CastMemberExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.active

        ws.append([
            'Nombre',
            'Apellido',
            'RUT',
            'Sexo',
            'Email',
            'Teléfono',
            'Fecha de nacimiento',
            'Edad',
            'Instrumento',
            'Orquesta',
        ])

        cast_members = self.filter_results(request)

        for cast_member in cast_members:
            try:
                orchestra = cast_member.orchestra.name
            except:
                orchestra = ''

            try:
                first_name = cast_member.personal_information.first_name
            except:
                first_name = ''

            try:
                last_name = cast_member.personal_information.last_name
            except:
                last_name = ''

            try:
                social_id = cast_member.personal_information.social_id
            except:
                social_id = ''

            try:
                gender = cast_member.personal_information.gender
            except:
                gender = ''

            try:
                email = cast_member.personal_information.email
            except:
                email = ''

            try:
                phone_number_mobile = cast_member.personal_information.phone_number_mobile
            except:
                phone_number_mobile = ''

            try:
                birth_date = cast_member.personal_information.birth_date
            except:
                birth_date = ''

            try:
                age = timezone.now().year - birth_date.year
            except:
                age = ''

            ws.append([
                first_name,
                last_name,
                social_id,
                gender,
                email,
                phone_number_mobile,
                birth_date,
                age,
                cast_member.instrument,
                orchestra,
            ])


        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content_type=EXCEL_MIME)
        response['Content-Disposition'] = 'attachment; filename={}'.format(FILENAME)
        response.write(stream)

        return response


    def filter_results(self, request):
        queryset = CastMember.objects.all()

        first_name = request.GET.get('first_name', None)
        last_name = request.GET.get('last_name', None)
        social_id = request.GET.get('social_id', None)
        gender = request.GET.get('gender', None)
        email = request.GET.get('email', None)
        phone_number_mobile = request.GET.get('phone_number_mobile', None)
        region = request.GET.get('region', None)
        province = request.GET.get('province', None)
        area = request.GET.get('area', None)
        instrument = request.GET.get('instrument', None)

        if first_name is not None:
            queryset = queryset.filter(
                personal_information__first_name__icontains=first_name,
            )

        if last_name is not None:
            queryset = queryset.filter(
                personal_information__last_name__icontains=last_name,
            )

        if social_id is not None:
            queryset = queryset.filter(
                personal_information__social_id__icontains=social_id,
            )

        if gender is not None:
            queryset = queryset.filter(
                personal_information__gender=gender,
            )

        if email is not None:
            queryset = queryset.filter(
                personal_information__email__icontains=email,
            )

        if phone_number_mobile is not None:
            queryset = queryset.filter(
                personal_information__phone_number_mobile__icontains=phone_number_mobile,
            )

        if region is not None:
            queryset = queryset.filter(
                orchestra__area__province__region_id=int(region),
            )

        if province is not None:
            queryset = queryset.filter(
                orchestra__area__province__id=int(province),
            )

        if area is not None:
            queryset = queryset.filter(
                orchestra__area__id=int(area),
            )

        if instrument is not None:
            queryset = queryset.filter(
                instrument__icontains=instrument,
            )

        return queryset
