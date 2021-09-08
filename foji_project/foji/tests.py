# -*- coding: utf-8 -*-

from django.test import TestCase

from openpyxl import load_workbook

from .factories.orchestra import OrchestraFactory

from .api.members_excel import (
    DirectorForm,
    InstructorForm,
    CastMemberForm,
    get_row_data,
    create_director_data,
    create_instructor_data,
    create_cast_member_data,
    create_orchestra_data,
)


class MembersExcelErrorsTestCase(TestCase):
    def test_error_contains_row_key(self):
        pass

    def test_error_contains_error_key(self):
        pass

    def test_error_row_key_is_number(self):
        pass

    def test_error_error_key_is_list(self):
        pass


class MembersExcelOrchestraTestCase(TestCase):
    def test_orchestra_upload_ok_contains_three_keys(self):
        workbook = load_workbook(filename='test_excel_upload_ok.xlsx')
        orchestra = OrchestraFactory()

        errors = create_orchestra_data(orchestra, workbook)

        self.assertTrue('director' in errors)
        self.assertTrue('instructors' in errors)
        self.assertTrue('cast_members' in errors)

    def test_orchestra_upload_bad_contains_three_keys(self):
        workbook = load_workbook(filename='test_excel_upload_bad.xlsx')
        orchestra = OrchestraFactory()

        errors = create_orchestra_data(orchestra, workbook)

        self.assertTrue('director' in errors)
        self.assertTrue('instructors' in errors)
        self.assertTrue('cast_members' in errors)

    def test_orchestra_upload_ok(self):
        workbook = load_workbook(filename='test_excel_upload_ok.xlsx')
        orchestra = OrchestraFactory()

        errors = create_orchestra_data(orchestra, workbook)

        director_errors = errors.get('director', [])
        instructors_errors = errors.get('instructors', [])
        cast_members_errors = errors.get('cast_members', [])

        self.assertTrue(len(director_errors) == 0)
        self.assertTrue(len(instructors_errors) == 0)
        self.assertTrue(len(cast_members_errors) == 0)

    def test_orchestra_upload_bad(self):
        workbook = load_workbook(filename='test_excel_upload_bad.xlsx')
        orchestra = OrchestraFactory()

        errors = create_orchestra_data(orchestra, workbook)

        director_errors = errors.get('director', [])
        instructors_errors = errors.get('instructors', [])
        cast_members_errors = errors.get('cast_members', [])

        test = len(director_errors) > 0 or len(instructors_errors) > 0 \
                or len(cast_members_errors) > 0

        self.assertTrue(test)


class MembersExcelCastMemberTestCase(TestCase):
    def test_cast_member_form_accepts_correct_rows(self):
        workbook = load_workbook(filename='test_excel_upload_ok.xlsx')

        all_data = get_row_data(workbook, 'Elenco', row=2)
        all_data_form = CastMemberForm(all_data)
        valid = all_data_form.is_valid()

        if not valid:
            print(all_data)
            print(all_data_form.errors)
            self.assertTrue(valid)

        optional_birth_date = get_row_data(workbook, 'Elenco', row=3)
        optional_birth_date_form = CastMemberForm(optional_birth_date)
        valid = optional_birth_date_form.is_valid()

        if not valid:
            print(optional_birth_date)
            print(optional_birth_date_form.errors)
            self.assertTrue(valid)

        optional_email = get_row_data(workbook, 'Elenco', row=4)
        optional_email_form = CastMemberForm(optional_email)
        valid = optional_email_form.is_valid()

        if not valid:
            print(optional_email)
            print(optional_email_form.errors)
            self.assertTrue(valid)

        optional_phone_number = get_row_data(workbook, 'Elenco', row=5)
        optional_phone_number_form = CastMemberForm(optional_phone_number)
        valid = optional_phone_number_form.is_valid()

        if not valid:
            print(optional_phone_number)
            print(optional_phone_number_form.errors)
            self.assertTrue(valid)

        optional_social_id = get_row_data(workbook, 'Elenco', row=6)
        optional_social_id_form = CastMemberForm(optional_social_id)
        valid = optional_social_id_form.is_valid()

        if not valid:
            print(optional_social_id)
            print(optional_social_id_form.errors)
            self.assertTrue(valid)

        missing_optionals = get_row_data(workbook, 'Elenco', row=7)
        missing_optionals_form = CastMemberForm(missing_optionals)
        valid = missing_optionals_form.is_valid()

        if not valid:
            print(missing_optionals)
            print(missing_optionals_form.errors)
            self.assertTrue(valid)

    def test_cast_member_form_rejects_incorrect_rows(self):
        workbook = load_workbook(filename='test_excel_upload_bad.xlsx')

        missing_first_name = get_row_data(workbook, 'Elenco', row=2)
        missing_first_name_form = CastMemberForm(missing_first_name)
        valid = missing_first_name_form.is_valid()

        if valid:
            print(missing_first_name)
            print(missing_first_name_form.errors)
            self.assertFalse(valid)

        missing_last_name = get_row_data(workbook, 'Elenco', row=3)
        missing_last_name_form = CastMemberForm(missing_last_name)
        valid = missing_last_name_form.is_valid()

        if valid:
            print(missing_last_name)
            print(missing_last_name_form.errors)
            self.assertFalse(valid)

        missing_instrument = get_row_data(workbook, 'Elenco', row=4)
        missing_instrument_form = CastMemberForm(missing_instrument)
        valid = missing_instrument_form.is_valid()

        if valid:
            print(missing_instrument)
            print(missing_instrument_form.errors)
            self.assertFalse(valid)

        missing_gender = get_row_data(workbook, 'Elenco', row=5)
        missing_gender_form = CastMemberForm(missing_gender)
        valid = missing_gender_form.is_valid()

        if valid:
            print(missing_gender)
            print(missing_gender_form.errors)
            self.assertFalse(valid)

    def test_excel_upload_ok(self):
        workbook = load_workbook(filename='test_excel_upload_ok.xlsx')

        orchestra = OrchestraFactory()

        errors = create_cast_member_data(orchestra, workbook)

        print(errors)

        self.assertTrue(len(errors) == 0)

    def test_excel_upload_bad(self):
        workbook = load_workbook(filename='test_excel_upload_bad.xlsx')
        orchestra = OrchestraFactory()

        errors = create_cast_member_data(orchestra, workbook)

        self.assertTrue(len(errors) != 0)


class MembersExcelInstructorTestCase(TestCase):
    def test_instructor_form_accepts_correct_rows(self):
        workbook = load_workbook(filename='test_excel_upload_ok.xlsx')

        all_data = get_row_data(workbook, 'Instructores', row=2)
        all_data_form = InstructorForm(all_data)
        valid = all_data_form.is_valid()

        if not valid:
            print(all_data)
            print(all_data_form.errors)
            self.assertTrue(valid)

        optional_birth_date = get_row_data(workbook, 'Instructores', row=3)
        optional_birth_date_form = InstructorForm(optional_birth_date)
        valid = optional_birth_date_form.is_valid()

        if not valid:
            print(optional_birth_date)
            print(optional_birth_date_form.errors)
            self.assertTrue(valid)

        optional_phone_number = get_row_data(workbook, 'Instructores', row=4)
        optional_phone_number_form = InstructorForm(optional_phone_number)
        valid = optional_phone_number_form.is_valid()

        if not valid:
            print(optional_phone_number)
            print(optional_phone_number_form.errors)
            self.assertTrue(valid)

        optional_social_id = get_row_data(workbook, 'Instructores', row=5)
        optional_social_id_form = InstructorForm(optional_social_id)
        valid = optional_social_id_form.is_valid()

        if not valid:
            print(optional_social_id)
            print(optional_social_id_form.errors)
            self.assertTrue(valid)

        missing_optionals = get_row_data(workbook, 'Instructores', row=6)
        missing_optionals_form = InstructorForm(missing_optionals)
        valid = missing_optionals_form.is_valid()

        if not valid:
            print(missing_optionals)
            print(missing_optionals_form.errors)
            self.assertTrue(valid)

    def test_instructor_form_rejects_incorrect_rows(self):
        workbook = load_workbook(filename='test_excel_upload_bad.xlsx')

        missing_first_name = get_row_data(workbook, 'Instructores', row=2)
        missing_first_name_form = InstructorForm(missing_first_name)
        valid = missing_first_name_form.is_valid()

        if valid:
            print(missing_first_name)
            print(missing_first_name_form.errors)
            self.assertFalse(valid)

        missing_last_name = get_row_data(workbook, 'Instructores', row=3)
        missing_last_name_form = InstructorForm(missing_last_name)
        valid = missing_last_name_form.is_valid()

        if valid:
            print(missing_last_name)
            print(missing_last_name_form.errors)
            self.assertFalse(valid)

        missing_instrument = get_row_data(workbook, 'Instructores', row=4)
        missing_instrument_form = InstructorForm(missing_instrument)
        valid = missing_instrument_form.is_valid()

        if valid:
            print(missing_instrument)
            print(missing_instrument_form.errors)
            self.assertFalse(valid)

        missing_gender = get_row_data(workbook, 'Instructores', row=5)
        missing_gender_form = InstructorForm(missing_gender)
        valid = missing_gender_form.is_valid()

        if valid:
            print(missing_gender)
            print(missing_gender_form.errors)
            self.assertFalse(valid)

        missing_email = get_row_data(workbook, 'Instructores', row=6)
        missing_email_form = InstructorForm(missing_email)
        valid = missing_email_form.is_valid()

        if valid:
            print(missing_email)
            print(missing_email_form.errors)
            self.assertFalse(valid)

    def test_excel_upload_ok(self):
        workbook = load_workbook(filename='test_excel_upload_ok.xlsx')

        orchestra = OrchestraFactory()

        errors = create_instructor_data(orchestra, workbook)

        print(errors)

        self.assertTrue(len(errors) == 0)

    def test_excel_upload_bad(self):
        workbook = load_workbook(filename='test_excel_upload_bad.xlsx')
        orchestra = OrchestraFactory()

        errors = create_instructor_data(orchestra, workbook)

        self.assertTrue(len(errors) != 0)


class MembersExcelDirectorTestCase(TestCase):
    def test_excel_load_and_print_row(self):
        workbook = load_workbook(filename='test_excel_upload_ok.xlsx')

        data = get_row_data(workbook, 'Director')

        print(data)

    def test_director_form_accepts_correct_rows(self):
        workbook = load_workbook(filename='test_excel_upload_ok.xlsx')

        all_data = get_row_data(workbook, 'Director', row=2)
        all_data_form = DirectorForm(all_data)
        valid = all_data_form.is_valid()

        if not valid:
            print(all_data)
            print(all_data_form.errors)
            self.assertTrue(valid)

        optional_gender = get_row_data(workbook, 'Director', row=3)
        optional_gender_form = DirectorForm(optional_gender)
        valid = optional_gender_form.is_valid()

        if not valid:
            print(optional_gender)
            print(optional_gender_form.errors)
            self.assertTrue(valid)

        optional_instrument = get_row_data(workbook, 'Director', row=4)
        optional_instrument_form = DirectorForm(optional_instrument)
        valid = optional_instrument_form.is_valid()

        if not valid:
            print(optional_instrument)
            print(optional_instrument_form.errors)
            self.assertTrue(valid)

        optional_phone_number = get_row_data(workbook, 'Director', row=5)
        optional_phone_number_form = DirectorForm(optional_phone_number)
        valid = optional_phone_number_form.is_valid()

        if not valid:
            print(optional_phone_number)
            print(optional_phone_number_form.errors)
            self.assertTrue(valid)

        optional_social_id = get_row_data(workbook, 'Director', row=6)
        optional_social_id_form = DirectorForm(optional_social_id)
        valid = optional_social_id_form.is_valid()

        if not valid:
            print(optional_social_id)
            print(optional_social_id_form.errors)
            self.assertTrue(valid)

        missing_optionals = get_row_data(workbook, 'Director', row=7)
        missing_optionals_form = DirectorForm(missing_optionals)
        valid = missing_optionals_form.is_valid()

        if not valid:
            print(missing_optionals)
            print(missing_optionals_form.errors)
            self.assertTrue(valid)

    def test_director_form_rejects_incorrect_rows(self):
        workbook = load_workbook(filename='test_excel_upload_bad.xlsx')

        missing_first_name = get_row_data(workbook, 'Director', row=2)
        missing_first_name_form = DirectorForm(missing_first_name)
        valid = missing_first_name_form.is_valid()

        if valid:
            print(missing_first_name)
            print(missing_first_name_form.errors)
            self.assertFalse(valid)

        missing_last_name = get_row_data(workbook, 'Director', row=3)
        missing_last_name_form = DirectorForm(missing_last_name)
        valid = missing_last_name_form.is_valid()

        if valid:
            print(missing_last_name)
            print(missing_last_name_form.errors)
            self.assertFalse(valid)

        missing_birth_date = get_row_data(workbook, 'Director', row=4)
        missing_birth_date_form = DirectorForm(missing_birth_date)
        valid = missing_birth_date_form.is_valid()

        if valid:
            print(missing_birth_date)
            print(missing_birth_date_form.errors)
            self.assertFalse(valid)

        missing_email = get_row_data(workbook, 'Director', row=5)
        missing_email_form = DirectorForm(missing_email)
        valid = missing_email_form.is_valid()

        if valid:
            print(missing_email)
            print(missing_email_form.errors)
            self.assertFalse(valid)

    def test_excel_upload_ok(self):
        workbook = load_workbook(filename='test_excel_upload_ok.xlsx')

        orchestra = OrchestraFactory()

        errors = create_director_data(orchestra, workbook)

        print(errors)

        self.assertTrue(len(errors) == 0)

    def test_excel_upload_bad(self):
        workbook = load_workbook(filename='test_excel_upload_bad.xlsx')
        orchestra = OrchestraFactory()

        errors = create_director_data(orchestra, workbook)

        self.assertTrue(len(errors) != 0)
