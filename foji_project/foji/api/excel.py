# -*- coding: utf-8 -*-

import re

from django import forms
from django.db import transaction

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework import serializers

from openpyxl import Workbook, load_workbook

from ..models.personal_information import PersonalInformation
from ..models.orchestra import Orchestra
from ..models.director import Director
from ..models.instructor import Instructor
from ..models.cast_member import CastMember


EXCEL_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

CAST_MEMBER_HEADER_CELLS = (
    'Nombre',
    'Apellido',
    'Instrumento',
    'Genero',
    'Fecha de nacimiento',
    'Email',
    'Teléfono',
    'RUT',
)

FILENAME = 'integrantes.xlsx'


class DirectorForm(forms.Form):
    first_name = forms.CharField()
    last_name = forms.CharField()
    gender = forms.CharField(required=False)
    birth_date= forms.DateField(required=False)
    email = forms.EmailField()
    phone_number_mobile = forms.CharField(required=False)
    social_id = forms.CharField(required=False)


class InstructorForm(forms.Form):
    first_name = forms.CharField()
    last_name = forms.CharField()
    gender = forms.CharField(required=False)
    birth_date= forms.DateField(required=False)
    email = forms.EmailField(required=False)
    phone_number_mobile = forms.CharField(required=False)
    social_id = forms.CharField(required=False)


class CastMemberForm(forms.Form):
    first_name = forms.CharField()
    last_name = forms.CharField()
    gender = forms.CharField()
    birth_date= forms.DateField(required=False)
    email = forms.EmailField(required=False)
    phone_number_mobile = forms.CharField(required=False)
    social_id = forms.CharField(required=False)


def create_row(member):
    try:
        first_name = member.personal_information.first_name
    except:
        first_name = None

    try:
        last_name = member.personal_information.last_name
    except:
        last_name = None

    try:
        instrument = member.instrument
    except:
        instrument = None

    try:
        gender = member.personal_information.gender
    except:
        gender = None

    try:
        birth_date = member.personal_information.birth_date
    except:
        birth_date = None

    try:
        email = member.personal_information.email
    except:
        email = None

    try:
        phone_number_mobile = str(member.personal_information.phone_number_mobile)
        phone_number_mobile = re.sub(r'\.0*$', '', phone_number_mobile)
    except:
        phone_number_mobile = str(member.personal_information.phone_number_mobile)

    try:
        social_id = str(member.personal_information.social_id)
        social_id = re.sub(r'\.0*$', '', social_id)
    except:
        social_id = str(member.personal_information.social_id)


    result = [
        first_name,
        last_name,
        instrument,
        gender,
        birth_date,
        email,
        phone_number_mobile,
        social_id,
    ]

    return result


class CastMembersExcelUploadAPI(APIView):
    """
    Handles the upload and then creation of cast members for an orchestra.
    """

    parser_classes = (MultiPartParser,)

    def post(self, request, pk=None):
        errors = dict()

        orchestra = Orchestra.objects.get(pk=pk)
        file_object = request.data.get('file')

        wb = load_workbook(file_object)

        sheet_director = wb['Director']
        sheet_instructors = wb['Instructores']
        sheet_cast = wb['Elenco']

        self.create_directors(sheet_director, orchestra, errors)
        self.create_instructors(sheet_instructors, orchestra, errors)
        self.create_cast(sheet_cast, orchestra, errors)

        orchestra.save()

        return Response(errors)

    def create_directors(self, sheet, orchestra, errors):
        data = {}

        try:
            data['first_name'] = str(sheet[2][0].value)
        except:
            data['first_name'] = None

        try:
            data['last_name'] = str(sheet[2][1].value)
        except:
            data['last_name'] = None

        try:
            data['instrument'] = str(sheet[2][2].value)
        except:
            data['instrument'] = None

        try:
            data['gender'] = str(sheet[2][3].value)
        except:
            data['gender'] = None

        try:
            data['birth_date'] = str(sheet[2][4].value)
        except:
            data['birth_date'] = None

        try:
            data['email'] = str(sheet[2][5].value)
        except:
            data['email'] = None

        try:
            pre = str(sheet[2][6].value)
            pre = re.sub(r'\.0*$', '', pre)
            data['phone_number_mobile'] = pre
        except:
            data['phone_number_mobile'] = str(sheet[2][6].value)

        try:
            pre = str(sheet[2][7].value)
            pre = re.sub(r'\.0*$', '', pre)
            data['social_id'] = pre
        except:
            data['social_id'] = str(sheet[2][7].value)


        form = DirectorForm(data)
        print(data)

        if form.is_valid():
            try:
                with transaction.atomic():
                    data = dict(form.cleaned_data)
                    personal_information = PersonalInformation.objects.create(
                        first_name=data['first_name'],
                        last_name=data['last_name'],
                        gender=data['gender'],
                        phone_number_mobile=data['phone_number_mobile'],
                        birth_date=data['birth_date'],
                        email=data['email'],
                        social_id=data['social_id'],
                    )

                    director = Director.objects.create(
                        personal_information=personal_information,
                        instrument=data['instrument'],
                    )

                    try:
                        orchestra.director.delete()
                    except:
                        pass

                    orchestra.director = director
                    orchestra.save()

            except Exception as e:
                if 'global' not in errors:
                    errors['global'] = []
                errors['global'].append('Error al crear director.')
                print(e)
        else:
            errors['director'] = form.errors


    def create_instructors(self, sheet, orchestra, errors):
        orchestra.instructors.all().delete()
        for row in sheet.iter_rows(min_row=2):
            data = {}

            try:
                data['first_name'] = row[0].value
            except:
                data['first_name'] = None

            try:
                data['last_name'] = row[1].value
            except:
                data['last_name'] = None

            try:
                data['instrument'] = row[2].value
            except:
                data['instrument'] = None

            try:
                data['gender'] = row[3].value
            except:
                data['gender'] = None

            try:
                data['birth_date'] = str(row[4].value)
            except:
                data['birth_date'] = None

            try:
                data['email'] = row[5].value
            except:
                data['email'] = None

            try:
                pre = str(row[6].value)
                pre = re.sub(r'\.0*$', '', pre)
                data['phone_number_mobile'] = pre
            except:
                data['phone_number_mobile'] = str(row[6].value)

            try:
                pre = str(row[7].value)
                pre = re.sub(r'\.0*$', '', pre)
                data['social_id'] = pre
            except:
                data['social_id'] = str(row[7].value)


            form = InstructorForm(data)

            if form.is_valid():
                try:
                    data = dict(form.cleaned_data)
                    personal_information = PersonalInformation.objects.create(
                        first_name=data['first_name'],
                        last_name=data['last_name'],
                        gender=data['gender'],
                        phone_number_mobile=data['phone_number_mobile'],
                        birth_date=data['birth_date'],
                        email=data['email'],
                        social_id=data['social_id'],
                    )

                    instructor = Instructor.objects.create(
                        personal_information=personal_information,
                        orchestra=orchestra,
                        instrument=data['instrument'],
                    )
                except Exception as e:
                    print(e)
                    continue

    def create_cast(self, sheet, orchestra, errors):
        print(sheet.max_row)
        orchestra.cast_members.all().delete()
        for row in sheet.iter_rows(min_row=2):
            data = {}

            try:
                pre = str(row[6].value)
                pre = re.sub(r'\.0*$', '', pre)
                data['phone_number_mobile'] = pre
            except:
                data['phone_number_mobile'] = str(row[6].value)

            try:
                pre = str(row[7].value)
                pre = re.sub(r'\.0*$', '', pre)
                data['social_id'] = pre
            except:
                data['social_id'] = str(row[7].value)

            try:
                data['first_name'] = row[0].value
                data['last_name'] = row[1].value
                data['instrument'] = row[2].value
                data['gender'] = row[3].value
                data['birth_date'] = str(row[4].value)
                data['email'] = row[5].value
            except:
                continue

            form = CastMemberForm(data)

            if form.is_valid():
                try:
                    data = dict(form.cleaned_data)
                    personal_information = PersonalInformation.objects.create(
                        first_name=data['first_name'],
                        last_name=data['last_name'],
                        gender=data['gender'],
                        phone_number_mobile=data['phone_number_mobile'],
                        birth_date=data['birth_date'],
                        email=data['email'],
                        social_id=data['social_id'],
                    )

                    cast_member = CastMember.objects.create(
                            personal_information=personal_information,
                            orchestra=orchestra,
                            instrument=data['instrument'],
                    )
                except Exception as e:
                    print(e)
                    continue

def create_row(member):
    result = [
        member.personal_information.first_name,
        member.personal_information.last_name,
        None,
        member.personal_information.gender,
        member.personal_information.birth_date,
        member.personal_information.email,
        member.personal_information.phone_number_mobile,
        member.personal_information.social_id,
    ]

    return result
