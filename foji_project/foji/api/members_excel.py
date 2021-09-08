# -*- coding: utf-8 -*-

from django import forms
from django.db import transaction

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework import serializers

from openpyxl import Workbook, load_workbook

from ..models.orchestra import Orchestra
from ..models.director import Director
from ..models.instructor import Instructor
from ..models.cast_member import CastMember
from ..models.personal_information import PersonalInformation


def excel_error(row, error):
    """Returns a dictionary representing the row and form errors."""
    return {
        'row': row,
        'error': error,
    }


class DirectorForm(forms.Form):
    first_name = forms.CharField()
    last_name = forms.CharField()
    instrument = forms.CharField(required=False)
    gender = forms.CharField(required=False)
    birth_date= forms.DateField(
        input_formats=['%d/%m/%Y'],
    )
    email = forms.CharField()
    phone_number_mobile = forms.CharField(required=False)
    social_id = forms.CharField(required=False)


class InstructorForm(forms.Form):
    first_name = forms.CharField()
    last_name = forms.CharField()
    instrument = forms.CharField()
    gender = forms.CharField()
    birth_date= forms.DateField(
        input_formats=['%d/%m/%Y'],
        required=False,
    )
    email = forms.CharField()
    phone_number_mobile = forms.CharField(required=False)
    social_id = forms.CharField(required=False)


class CastMemberForm(forms.Form):
    first_name = forms.CharField()
    last_name = forms.CharField()
    instrument = forms.CharField()
    gender = forms.CharField()
    birth_date= forms.DateField(
        input_formats=['%d/%m/%Y'],
        required=False,
    )
    email = forms.CharField(required=False)
    phone_number_mobile = forms.CharField(required=False)
    social_id = forms.CharField(required=False)


def empty_data(data):
    """True if all data is ``blank'' (None or empty string)"""
    for key, value in data.items():
        spaces = False

        try:
            spaces = value.isspace()
        except:
            pass

        if spaces:
            continue

        if value:
            return False

    return True


def get_row_data(workbook, sheet_name='Director', row=2):
    """
    Gets row data from a specific workbook, sheet and row.
    """

    sheet = workbook[sheet_name]

    data = {}

    data['first_name'] = sheet[row][0].value
    data['last_name'] = sheet[row][1].value
    data['instrument'] = sheet[row][2].value

    data['gender'] = sheet[row][3].value
    if data['gender'] not in ('M', 'F'):
        data['gender'] = None

    data['birth_date'] = sheet[row][4].value
    data['email'] = sheet[row][5].value
    data['phone_number_mobile'] = sheet[row][6].value
    data['social_id'] = sheet[row][7].value

    return data


def create_director_data(orchestra, workbook, sheet_name='Director'):
    """
    Creates a director from a ``directors'' sheet in a workbook.
    """

    # We return errors generated in the sheet.
    errors = []

    # Get data from the workbook.
    data = get_row_data(workbook, sheet_name, 2) # Second row.

    form = DirectorForm(data)

    if form.is_valid():
        try:
            # We try to get a director instance for the orchestra. If it
            # doesn't exist we create one.
            director = orchestra.director

            if director is None:
                director = Director.directors.create()

            personal_information = director.personal_information

            with transaction.atomic():
                personal_information.first_name = form.cleaned_data.get('first_name')
                personal_information.last_name = form.cleaned_data.get('last_name') 
                director.instrument = form.cleaned_data.get('instrument')
                personal_information.gender = form.cleaned_data.get('gender')
                personal_information.birth_date = form.cleaned_data.get('birth_date')
                personal_information.email = form.cleaned_data.get('email')
                personal_information.phone_number_mobile = form.cleaned_data.get('phone_number_mobile')
                personal_information.social_id = form.cleaned_data.get('social_id')

                personal_information.save()
                director.save()

                orchestra.director = director
                orchestra.save()

        except Exception as e:
            print(e)
            error = excel_error(
                0,
                'No se pudo crear director.'
            )

            errors.append(error)
    else:
        error = excel_error(
            2,
            form.errors,
        )

        errors.append(error)

    return errors


def create_instructor_data(orchestra, workbook, sheet_name='Instructores'):
    """
    Creates instructors from an ``instructors'' sheet in a workbook.
    """

    # We return errors generated in the sheet.
    errors = []

    # Get data from the workbook.

    sheet = workbook[sheet_name]

    try:
        with transaction.atomic():
            # Delete all instructors.
            orchestra.instructors.all().delete()

            for row in sheet.iter_rows(min_row=2):
                row_number = row[0].row
                print(row_number)

                data = get_row_data(workbook, sheet_name, row_number) # Second row.

                # Skip if empty
                if empty_data(data):
                    continue

                print(data)

                form = InstructorForm(data)

                if form.is_valid():
                    personal_information = PersonalInformation.objects.create(
                        first_name=form.cleaned_data.get('first_name'),
                        last_name=form.cleaned_data.get('last_name'),
                        gender=form.cleaned_data.get('gender'),
                        phone_number_mobile=form.cleaned_data.get('phone_number_mobile'),
                        birth_date=form.cleaned_data.get('birth_date'),
                        email=form.cleaned_data.get('email'),
                        social_id=form.cleaned_data.get('social_id'),
                    )

                    Instructor.objects.create(
                        personal_information=personal_information,
                        orchestra=orchestra,
                        instrument=form.cleaned_data.get('instrument')
                    )
                else:
                    error = excel_error(
                        row_number,
                        form.errors,
                    )
                    errors.append(error)
    except Exception as e:
        print(e)
        error = excel_error(
            0,
            str('No se pudo crear instructores'),
        )
        errors.append(error)

    return errors


def create_cast_member_data(orchestra, workbook, sheet_name='Elenco'):
    """
    Creates instructors from a ``cast members'' sheet in a workbook.
    """

    # We return errors generated in the sheet.
    errors = []

    # Get data from the workbook.

    sheet = workbook[sheet_name]

    try:
        with transaction.atomic():
            # Delete all cast members.
            orchestra.cast_members.all().delete()

            for row in sheet.iter_rows(min_row=2):
                row_number = row[0].row
                print(row_number)

                data = get_row_data(workbook, sheet_name, row_number) # Second row.

                # Skip if empty
                if empty_data(data):
                    continue

                print(data)

                form = CastMemberForm(data)

                if form.is_valid():
                    personal_information = PersonalInformation.objects.create(
                        first_name=form.cleaned_data.get('first_name'),
                        last_name=form.cleaned_data.get('last_name'),
                        gender=form.cleaned_data.get('gender'),
                        phone_number_mobile=form.cleaned_data.get('phone_number_mobile'),
                        birth_date=form.cleaned_data.get('birth_date'),
                        email=form.cleaned_data.get('email'),
                        social_id=form.cleaned_data.get('social_id'),
                    )

                    CastMember.objects.create(
                        personal_information=personal_information,
                        orchestra=orchestra,
                        instrument=form.cleaned_data.get('instrument')
                    )
                else:
                    error = excel_error(
                        row_number,
                        form.errors,
                    )
                    errors.append(error)
    except Exception as e:
        print(e)
        error = excel_error(
            0,
            'No se pudo crear elenco'
        )
        errors.append(error)

    return errors


def create_orchestra_data(orchestra, workbook):
    director_errors = []
    instructor_errors = []
    cast_member_errors = []

    try:
        with transaction.atomic():
            director_errors = create_director_data(orchestra, workbook)
            if len(director_errors) > 0:
                raise

            instructor_errors = create_instructor_data(orchestra, workbook)
            if len(instructor_errors) > 0:
                raise

            cast_member_errors = create_cast_member_data(orchestra, workbook)
            if len(cast_member_errors) > 0:
                raise
    except Exception as e:
        print(e)

    return {
        'director': director_errors,
        'instructors': instructor_errors,
        'cast_members': cast_member_errors,
    }


class CastMembersExcelUploadAPI(APIView):
    """
    Handles the upload and then creation of cast members for an orchestra.
    """

    parser_classes = (MultiPartParser,)

    def post(self, request, pk=None):
        orchestra = Orchestra.objects.get(pk=pk)
        file_object = request.data.get('file')

        workbook = load_workbook(file_object)

        errors = create_orchestra_data(orchestra, workbook)

        return Response(errors)
