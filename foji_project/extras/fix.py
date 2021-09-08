# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook

from foji.models.coordinator import Coordinator

from .utils import (
    get_row_data,
    empty_cell_data,
    empty_data,
)

def empty_util(data):
    if data is None:
        return True

    if data == 'None':
        return True

    return empty_cell_data(data)


def run(row=2, debug=True):
    filename = input('Excel file (.xlsx): ')
    workbook = load_workbook(filename=filename)

    return fix_coordinators(workbook, min_row=row, debug=debug)


def fix_coordinators(workbook, min_row=2, debug=True):
    successes = []
    errors = []

    sheet = workbook.active

    for row in sheet.iter_rows(min_row=min_row):
        row_number = row[0].row

        data = get_row_data(workbook, row=row_number)
        if empty_data(data):
            continue

        success, error = fix_coordinator_phone_number_data(workbook, row_number, debug)

        if success:
            successes.append(error)
        else:
            errors.append(error)

    return successes, errors


def fix_coordinator_phone_number_data(workbook, row, debug=True):
    """
    Fix `phone_number_mobile' for coordinator if not present. Returns a tuple
    reporting the success and error if present.
    """

    # Get row data.
    data = get_row_data(workbook, row=row)
    print(data)
    coordinator_username = data.get('coordinator_email')
    phone_number = data.get('coordinator_phone_number_mobile')
    print('Phone number for {}: {}'.format(coordinator_username, phone_number))

    if not coordinator_username:
        return False, 'ERROR {}: No hay email de coordinador.'.format(row)

    # Find coordinator.
    try:
        coordinator = Coordinator.objects.get(user__username__iexact=coordinator_username)
    except Exception as e:
        return False, 'ERROR {}: No existe coordinador con email {}.\n{}'.format(
            row,
            coordinator_username,
            e,
        )

    old_phone_number = coordinator.personal_information.phone_number_mobile

    if old_phone_number is None:
        print('Old is None')
    else:
        print('Old is not None: {}'.format(old_phone_number))

    if not empty_util(old_phone_number):
        return True, 'SUCCESS {}: {}, keep old {}'.format(
            row,
            coordinator_username,
            old_phone_number,
        )

    if empty_cell_data(phone_number):
        return False, 'ERROR {}: {}, Telefono vacio.'.format(
            row,
            coordinator_username,
        )


    if not debug:
        personal_information = coordinator.personal_information

        try:
            personal_information.phone_number_mobile = phone_number
            personal_information.save()
        except Exception as e:
            return False, 'ERROR {}: {}'.format(
                row,
                e,
            )

    success_message = 'SUCCESS {}: {} number to {}.'.format(
        row,
        coordinator_username,
        phone_number,
    )

    if debug:
        print(success_message)

    return True, success_message
