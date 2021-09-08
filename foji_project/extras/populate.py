# -*- coding: utf-8 -*-

from django import forms
from django.db import transaction

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework import serializers

from openpyxl import Workbook, load_workbook

from ..models.orchestra import Orchestra
from ..models.director import Director
from ..models.instructor import Instructor
from ..models.cast_member import CastMember
from ..models.personal_information import PersonalInformation


class ExcelRowForm(forms.Form):
    orchestra_name = forms.CharField()
    creation_date= forms.IntegerField()
    city = forms.CharField()
    area_name = forms.CharField()
    orchestra_status = forms.CharField()
    orchestra_type = forms.CharField()
    institution_name = forms.CharField()
    coordinator_name = forms.CharField()
    coordinator_phone_number_mobile = forms.CharField()
    coordinator_email = forms.EmailField()
    director_name = forms.CharField()
    director_phone_number_mobile = forms.CharField()
    director_email = forms.EmailField()


def excel_error(row, error):
    """Returns a dictionary representing the row and form errors."""
    return {
        'row': row,
        'error': error,
    }


def empty_data(data):
    """True if all data is ``blank'' (None or empty string)"""
    for key, value in data.items():
        spaces = False

        try:
            spaces = value.isspace()
        except:
            pass

        if spaces:
            continue

        if value:
            return False

    return True


def get_row_data(workbook, sheet_name=None, row=2):
    """
    Gets row data from a specific workbook, sheet and row.
    """

    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    data = {}

    data['orchestra_name'] = sheet[row][0].value
    data['creation_date'] = sheet[row][1].value
    data['city'] = sheet[row][4].value
    data['area_name'] = sheet[row][5].value
    data['orchestra_status'] = sheet[row][6].value
    data['orchestra_type'] = sheet[row][7].value
    data['institution_name'] = sheet[row][9].value
    data['coordinator_name'] = sheet[row][10].value
    data['coordinator_phone_number_mobile'] = sheet[row][11].value
    data['coordinator_email'] = sheet[row][12].value
    data['director_name'] = sheet[row][13].value
    data['director_phone_number_mobile'] = sheet[row][14].value
    data['director_email'] = sheet[row][15].value

    return data



def create_orchestra(data, institution, coordinator, director):
    """Creates an orchestra."""

    errors = []

    try:
        area = Area.objects.get(name__iexact=data.get('area_name'))
    except:
        area = None

    orchestra = Orchestra.orchestras.create(
        name=data.get('orchestra_name'),
        institution=institution,
        coordinator=coordinator,
        director=director,
        institution=institution,
        orchestra_status=data.get('orchestra_status', 'paused'),
        orchestra_type=data.get('orchestra_type'),
        area=area,
        city=data.get('city'),
        creation_date=data.get('creation_date'),
    )

    if orchestra is None:
        errors.append('No se puedo crear orquesta')

    return orchestra, errors


def create_institution(data):
    pass


def create_coordinator(data):
    """Creates a coordinator."""

    errors = []

    first_name = data.get('coordinator_name')
    last_name = None

    coordinator = Coordinator.coordinators.create(
        email=data.get('coordinator_email'),
        first_name=first_name,
        last_name=last_name,
        phone_number_mobile=data.get('coordinator_phone_number_mobile'),
    )

    if coordinator is None:
        errors.append('No se puedo crear coordinador')

    return coordinator, errors


def create_director(data):
    """Creates a director."""

    errors = []

    first_name = data.get('director_name')
    last_name = None

    director = Director.directors.create(
        first_name=first_name,
        last_name=last_name,
        email=data.get('director_email'),
    )

    if director is None:
        errors.append('No se pudo crear director.')

    return director, errors


def create_row_data(data):
    """
    Creates the models for the orchestra, institution, coordinator and
    director.
    """
    errors = []

    with transaction.atomic():
        coordinator, coordinator_error = create_coordinator(data)

        if coordinator is None:

        director, director_error = create_director(data)
        institution, institution_error = create_institution(data)
        orchestra, orchestra_error = create_orchestra(data, institution, coordinator, director)

    return orchestra, errors


def create_all_row_data(workbook, sheet_name=None):
    for row in sheet.iter_rows(min_row=2):
        row_number = row[0].row


def create_director_data(orchestra, workbook, sheet_name='Director'):
    """
    Creates a director from a ``directors'' sheet in a workbook.
    """

    # We return errors generated in the sheet.
    errors = []

    # Get data from the workbook.
    data = get_row_data(workbook, sheet_name, 2) # Second row.

    form = DirectorForm(data)

    if form.is_valid():
        try:
            # We try to get a director instance for the orchestra. If it
            # doesn't exist we create one.
            director = orchestra.director

            if director is None:
                director = Director.directors.create()

            personal_information = director.personal_information

            with transaction.atomic():
                personal_information.first_name = form.cleaned_data.get('first_name')
                personal_information.last_name = form.cleaned_data.get('last_name') 
                director.instrument = form.cleaned_data.get('instrument')
                personal_information.gender = form.cleaned_data.get('gender')
                personal_information.birth_date = form.cleaned_data.get('birth_date')
                personal_information.email = form.cleaned_data.get('email')
                personal_information.phone_number_mobile = form.cleaned_data.get('phone_number_mobile')
                personal_information.social_id = form.cleaned_data.get('social_id')

                personal_information.save()
                director.save()

                orchestra.director = director
                orchestra.save()

        except Exception as e:
            print(e)
            error = excel_error(
                0,
                'No se pudo crear director.'
            )

            errors.append(error)
    else:
        error = excel_error(
            2,
            form.errors,
        )

        errors.append(error)

    return errors


def create_instructor_data(orchestra, workbook, sheet_name='Instructores'):
    """
    Creates instructors from an ``instructors'' sheet in a workbook.
    """

    # We return errors generated in the sheet.
    errors = []

    # Get data from the workbook.

    sheet = workbook[sheet_name]

    try:
        with transaction.atomic():
            # Delete all instructors.
            orchestra.instructors.all().delete()

            for row in sheet.iter_rows(min_row=2):
                row_number = row[0].row
                print(row_number)

                data = get_row_data(workbook, sheet_name, row_number) # Second row.

                # Skip if empty
                if empty_data(data):
                    continue

                print(data)

                form = InstructorForm(data)

                if form.is_valid():
                    personal_information = PersonalInformation.objects.create(
                        first_name=form.cleaned_data.get('first_name'),
                        last_name=form.cleaned_data.get('last_name'),
                        gender=form.cleaned_data.get('gender'),
                        phone_number_mobile=form.cleaned_data.get('phone_number_mobile'),
                        birth_date=form.cleaned_data.get('birth_date'),
                        email=form.cleaned_data.get('email'),
                        social_id=form.cleaned_data.get('social_id'),
                    )

                    Instructor.objects.create(
                        personal_information=personal_information,
                        orchestra=orchestra,
                        instrument=form.cleaned_data.get('instrument')
                    )
                else:
                    error = excel_error(
                        row_number,
                        form.errors,
                    )
                    errors.append(error)
    except Exception as e:
        print(e)
        error = excel_error(
            0,
            str('No se pudo crear instructores'),
        )
        errors.append(error)

    return errors


def create_cast_member_data(orchestra, workbook, sheet_name='Elenco'):
    """
    Creates instructors from a ``cast members'' sheet in a workbook.
    """

    # We return errors generated in the sheet.
    errors = []

    # Get data from the workbook.

    sheet = workbook[sheet_name]

    try:
        with transaction.atomic():
            # Delete all cast members.
            orchestra.cast_members.all().delete()

            for row in sheet.iter_rows(min_row=2):
                row_number = row[0].row
                print(row_number)

                data = get_row_data(workbook, sheet_name, row_number) # Second row.

                # Skip if empty
                if empty_data(data):
                    continue

                print(data)

                form = CastMemberForm(data)

                if form.is_valid():
                    personal_information = PersonalInformation.objects.create(
                        first_name=form.cleaned_data.get('first_name'),
                        last_name=form.cleaned_data.get('last_name'),
                        gender=form.cleaned_data.get('gender'),
                        phone_number_mobile=form.cleaned_data.get('phone_number_mobile'),
                        birth_date=form.cleaned_data.get('birth_date'),
                        email=form.cleaned_data.get('email'),
                        social_id=form.cleaned_data.get('social_id'),
                    )

                    CastMember.objects.create(
                        personal_information=personal_information,
                        orchestra=orchestra,
                        instrument=form.cleaned_data.get('instrument')
                    )
                else:
                    error = excel_error(
                        row_number,
                        form.errors,
                    )
                    errors.append(error)
    except Exception as e:
        print(e)
        error = excel_error(
            0,
            'No se pudo crear elenco'
        )
        errors.append(error)

    return errors


def create_orchestra_data(orchestra, workbook):
    director_errors = []
    instructor_errors = []
    cast_member_errors = []

    try:
        with transaction.atomic():
            director_errors = create_director_data(orchestra, workbook)
            if len(director_errors) > 0:
                raise

            instructor_errors = create_instructor_data(orchestra, workbook)
            if len(instructor_errors) > 0:
                raise

            cast_member_errors = create_cast_member_data(orchestra, workbook)
            if len(cast_member_errors) > 0:
                raise
    except Exception as e:
        print(e)

    return {
        'director': director_errors,
        'instructors': instructor_errors,
        'cast_members': cast_member_errors,
    }


class CastMembersExcelUploadAPI(APIView):
    """
    Handles the upload and then creation of cast members for an orchestra.
    """

    parser_classes = (MultiPartParser,)

    def post(self, request, pk=None):
        orchestra = Orchestra.objects.get(pk=pk)
        file_object = request.data.get('file')

        workbook = load_workbook(file_object)

        errors = create_orchestra_data(orchestra, workbook)

        return Response(errors)
