# -*- coding: utf-8 -*-

import datetime

from django.db import transaction

from openpyxl import Workbook, load_workbook

from foji.models.coordinator import Coordinator
from foji.models.orchestra import Orchestra
from foji.models.director import Director
from foji.models.institution_information import InstitutionInformation
from foji.models.location import Area

excel_file = open('base5ok.xlsx', 'rb')

wb = load_workbook(excel_file)
ws = wb.active

coordinators_created = []

for row in ws.iter_rows(min_row=2):
    try:
        with transaction.atomic():
            # Orchestra
            try:
                name = row[0].value.title()
            except:
                continue

            try:
                creation_year = int(row[1].value)
                creation_date = datetime.date(creation_year, 1, 1)
            except Exception as e:
                print('When looking at the year:')
                print(e)
                creation_date = None

            try:
                orchestra_status = row[6].value

                if orchestra_status == 'ACTIVA':
                    orchestra_status = 'active'
                elif orchestra_status == 'EN PAUSA':
                    orchestra_status = 'paused'
                else:
                    orchestra_status = 'inactive'
            except:
                orchestra_status = 'active'

            try:
                city = row[4].value.title()
            except:
                city = ''

            try:
                x = row[5].value
                print('Raw area name:', x, 'length:', len(x))
                area_name = row[5].value.title()
                print('Area:', area_name)
                area = Area.objects.get(name__iexact=area_name)
            except Exception as e:
                print('Error when getting area:', e)
                area = None

            try:
                orchestra_type = row[7].value.lower()
            except:
                orchestra_type = None

            orchestra = Orchestra(
                name=name,
                creation_date=creation_date,
                orchestra_status=orchestra_status,
                city=city,
                area=area,
                orchestra_type=orchestra_type,
            )

            # Director
            try:
                full_name = row[13].value.title()
            except:
                full_name = ''

            try:
                full_name_split = full_name.split(' ')
                first_name = full_name_split[0]
                last_name = ' '.join(full_name_split[1:])
            except:
                first_name = full_name
                last_name = None

            try:
                phone_number_mobile = str(row[14].value)
            except:
                phone_number_mobile = ''

            try:
                email = row[15].value.lower()
            except:
                email = None

            director = Director.directors.create(
                first_name=first_name,
                last_name=last_name,
                email=email,
            )

            # Institution
            try:
                name = row[9].value
            except:
                name = None

            institution_information = \
                InstitutionInformation.institution_information.create(
                    name=name,
                )

            # Coordinator
            try:
                full_name = row[10].value.title()
            except:
                full_name = ''

            try:
                full_name_split = full_name.split(' ')
                first_name = full_name_split[0]
                last_name = ' '.join(full_name_split[1:])
            except:
                first_name = full_name
                last_name = None

            try:
                phone_number_mobile = str(row[11].value)
                print(phone_number_mobile)
            except Exception as e:
                print('Phone shit')
                print(e)
                phone_number_mobile = ''

            try:
                email = row[12].value.lower()
                print('email:', email)
            except:
                print('There was no email for coordinator')
                continue

            try:
                coordinator = Coordinator.objects.get(user__username=email)
            except:
                coordinator = Coordinator.coordinators.create(
                    first_name=first_name,
                    last_name=last_name,
                    phone_number_mobile=phone_number_mobile,
                    role='Coordinador',
                    email=email,
                )

            if coordinator is None:
                print('Cannot create coordinator skipping orchestra.')
                continue

            orchestra.director = director
            orchestra.institution = institution_information
            orchestra.coordinator = coordinator
            orchestra.save()

    except Exception as e:
        print('While creating orchestra there was an error in the transaction.')
        print(e)
        continue

    print(orchestra)
    coordinators_created.append(orchestra.coordinator)


print(len(coordinators_created))
for c in coordinators_created:
    c.send_beta_email()
    print('======\n\n\nEnviando correo a:')
    print(c.user)
